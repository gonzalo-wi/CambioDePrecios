from django.shortcuts import render


import pandas as pd
import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from AppCambioPrecios.models import *
from AppCambioPrecios.forms import *
from AppCambioPrecios.urls import *
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from .sincronizacion import *
from django.db.models import Q

@login_required
def inicio(request):
    return render(request, "AppCambioPrecios/inicio.html")


@login_required
def listas(request):
    precios = Precio.objects.all()
    mensaje = ""

    if request.method == 'POST':
        for key in request.POST.keys():
            if key.startswith('precio_'):
                id_producto = key.split('_')[1]
                nuevo_precio = request.POST[key].strip()

                if nuevo_precio:
                    try:
                        nuevo_precio_float = float(nuevo_precio)
                        if nuevo_precio_float < 0:
                            messages.error(request, f"El precio ingresado para el producto {id_producto} no puede ser negativo.")
                            continue

                        precios = Precio.objects.filter(idProducto=id_producto)
                        for precio_obj in precios:
                            guardar_precio_anterior(precio_obj)
                            precio_obj.precio = nuevo_precio_float
                            precio_obj.save()

                    except ValueError:
                        messages.error(request, f"El precio ingresado para el producto {id_producto} no es válido.")

        messages.success(request, "Precios actualizados correctamente.")
        return redirect('listas')

    # Agrupar los precios por nombre de lista y agregar precios anteriores
    precios_por_hoja = {}
    for precio in precios:
        if precio.nombreDeLista not in precios_por_hoja:
            precios_por_hoja[precio.nombreDeLista] = []
        # Obtener el precio anterior
        precio_anterior = PrecioAntiguo.objects.filter(idListaPrecio=precio.idListaPrecio, idProducto=precio.idProducto).order_by('-id').first()
        precio.precio_anterior = precio_anterior.precio_anterior if precio_anterior else None
        precios_por_hoja[precio.nombreDeLista].append(precio)

    return render(request, 'AppCambioPrecios/listas.html', {'precios_por_hoja': precios_por_hoja, 'mensaje': mensaje})

@login_required
def subir_archivo(request):
    if request.method == 'POST':
        archivo_excel = request.FILES['archivo_excel']
        try:
            wb = openpyxl.load_workbook(archivo_excel)
            for hoja in wb.sheetnames:
                sheet = wb[hoja]  

                for fila in sheet.iter_rows(min_row=2, values_only=True):  
                    idListaPrecio, idProducto, descripcion, precio = fila

                    if idListaPrecio and idProducto and descripcion and precio is not None:
                        
                        precio_obj, created = Precio.objects.get_or_create(
                            nombreDeLista=hoja,
                            idListaPrecio=idListaPrecio,
                            idProducto=idProducto,
                            defaults={'descripcion': descripcion, 'precio': precio}
                        )

                        if not created:
                            guardar_precio_anterior(precio_obj)
                            precio_obj.descripcion = descripcion
                            precio_obj.precio = precio
                            precio_obj.save()
                            
            conn = obtener_conexion()
            if conn:
                cursor = conn.cursor()
                guardar_precios_clientes(cursor)
                conn.commit()
                cursor.close()
                conn.close()         
        except Exception as e:
            messages.error(request, f"Error al subir el archivo: {e}")
        else:
            messages.success(request, "Archivo subido y procesado correctamente.")
        return redirect('listas')

    return render(request, 'AppCambioPrecios/subir_archivo.html')
@login_required
def sincronizar_precios_view(request):
    if request.method == 'POST':
        resultado = sincronizar_precios()
        if "Error" in resultado:
            messages.error(request, resultado)
        else:
            messages.success(request, resultado)
        
        print(f"Mensaje: {resultado}")
        return redirect('sincronizar_precios')

    return render(request, 'AppCambioPrecios/sincronizar_precios.html')

@login_required
def restaurar_precios_view(request):
    if request.method == 'POST':
        resultado = restaurar_precios(request)
        if "Error" in resultado:
            messages.error(request, resultado)
        else:
            messages.success(request, resultado)
        
        print(f"Mensaje: {resultado}")
        return redirect('restaurar_precios')

    return render(request, 'AppCambioPrecios/sincronizar_precios.html')
@login_required
def ejecutar_primer_script_view(request):
    if request.method == 'POST':
        try:
            ejecutar_primer_script()
            messages.success(request, "Script SQL ejecutado correctamente.")
        except Exception as e:
            messages.error(request, f"Error al ejecutar el script SQL: {e}")
        return redirect('ejecutar_primer_script')
    return render(request, 'AppCambioPrecios/primerScript.html')

@login_required
def ejecutar_segundo_script_view(request):
    if request.method == 'POST':
        try:
            ejecutar_segundo_script()
            messages.success(request, "Script ejecutado exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al ejecutar el script: {e}")
        return redirect('ejecutar_segundo_script')  # Redirige a la vista deseada después de ejecutar el script

    return render(request, 'AppCambioPrecios/segundoScript.html')

@login_required
def buscar_listas_precios(request):
    precios = []
    mensaje = ""

    if request.method == 'POST':
        id_lista = request.POST.get('id_lista')
        id_producto = request.POST.get('id_producto')

        if id_lista and id_producto:
            precios = Precio.objects.filter(
                Q(idListaPrecio=id_lista) & Q(idProducto=id_producto)
            )
            if not precios:
                mensaje = "No se encontraron listas de precios con los criterios especificados."
        else:
            mensaje = "Por favor, ingrese ambos ID de lista y ID de producto."

    precios_anteriores = {precio.idProducto: PrecioAntiguo.objects.filter(idProducto=precio.idProducto).order_by('-id').first() for precio in precios}

    return render(request, 'AppCambioPrecios/listas.html', {'precios': precios, 'precios_anteriores': precios_anteriores, 'mensaje': mensaje})
